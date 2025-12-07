# This file contains dead code that was removed from the project.
# It is kept here for reference and potential future use.

# "Workout Scheduler" code from pyomo.py
from google.colab import auth
auth.authenticate_user()

import gspread
from google.auth import default
from datetime import datetime, timedelta

creds, _ = default()
gc = gspread.authorize(creds)

spreadsheet = gc.create('Workout_Scheduler_Template')
print(f"Created: {spreadsheet.url}")

# UserData sheet (first sheet)
user_sheet = spreadsheet.sheet1
user_sheet.update_title('UserData')
user_sheet.update(values=[['user_id', '1']], range_name='A1:B1')
user_sheet.format('A1:B1', {'textFormat': {'bold': True}})

# Schedule sheet
schedule_sheet = spreadsheet.add_worksheet('Schedule', rows=50, cols=5)
schedule_sheet.update(values=[['Date', 'Workout_Sheet_Name', 'Completed', 'Actual_Duration', 'Notes']], range_name='A1:E1')

today = datetime.now()
rows = [[(today + timedelta(days=i)).strftime('%Y-%m-%d'), f'Workout_Day_{i+1}', 'No', '', ''] for i in range(7)]
schedule_sheet.update(values=rows, range_name='A2:E8')

# Workout 1
w1 = spreadsheet.add_worksheet('Workout_Day_1', rows=15, cols=5)
w1.update(values=[['Interval_Order', 'Duration_Sec', 'Pace', 'Cadence', 'Type']], range_name='A1:E1')
w1.update(values=[[1,600,'6:00/km',85,'warmup'],[2,240,'5:00/km',90,'work'],[3,120,'6:30/km',80,'recovery'],[4,240,'5:00/km',90,'work'],[5,120,'6:30/km',80,'recovery'],[6,240,'5:00/km',90,'work'],[7,600,'6:00/km',85,'cooldown']], range_name='A2:E8')
w1.update(values=[['Location: track'],['Music: tempo_mix.mp3'],['Notes: Tempo intervals']], range_name='A10:A12')

# Workout 2
w2 = spreadsheet.add_worksheet('Workout_Day_2', rows=15, cols=5)
w2.update(values=[['Interval_Order', 'Duration_Sec', 'Pace', 'Cadence', 'Type']], range_name='A1:E1')
w2.update(values=[[1,300,'6:00/km',85,'warmup'],[2,1800,'6:00/km',85,'steady'],[3,300,'6:30/km',80,'cooldown']], range_name='A2:E4')
w2.update(values=[['Location: park'],['Music: easy_mix.mp3'],['Notes: Easy run']], range_name='A6:A8')

# Workout 3
w3 = spreadsheet.add_worksheet('Workout_Day_3', rows=15, cols=5)
w3.update(values=[['Interval_Order', 'Duration_Sec', 'Pace', 'Cadence', 'Type']], range_name='A1:E1')
w3.update(values=[[1,600,'6:00/km',85,'warmup'],[2,120,'5:30/km',88,'hill'],[3,120,'7:00/km',75,'recovery'],[4,120,'5:30/km',88,'hill'],[5,120,'7:00/km',75,'recovery'],[6,120,'5:30/km',88,'hill'],[7,600,'6:00/km',85,'cooldown']], range_name='A2:E8')
w3.update(values=[['Location: hills'],['Music: hill_mix.mp3'],['Notes: Hill repeats']], range_name='A10:A12')

print("✅ Done!")
print(f"📋 URL: {spreadsheet.url}")
print("\nUserData sheet created with user_id=1 (edit B1 to change)")

# Data transformation code from pyomo.py
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import date

# --- Assumes these are already loaded ---
# npt = pd.read_pickle(npt.pkl)
# chosenfoods1 = pd.read_excel(chosenfoods1.xlsx, index_col=0)

# --- Step 1: Set up food IDs ---
green_cabbage_id = '11109'
green_sauerkraut_id = '11439'
red_cabbage_id = '11112'
red_kraut_id = 'redkraut'

# --- Step 2: Create base from red cabbage ---
rk = npt.loc[red_cabbage_id].astype(float).copy()  # Full baseline nutrient row


# --- Step 3: Apply fermentation transformation to selected nutrients ---
transform_nutrients = ['203', '204', '208', '269', '212', '291', '401']

gc = npt.loc[green_cabbage_id].astype(float)
gs = npt.loc[green_sauerkraut_id].astype(float)

gc_safe = gc.replace(0, np.nan)
ratios = gs[transform_nutrients] / gc_safe[transform_nutrients]

# Apply transformation where possible
for nutr in transform_nutrients:
    if pd.notna(ratios[nutr]) and nutr in rk:
        rk[nutr] = round(rk[nutr] * ratios[nutr], 3)

# --- Step 4: Override sodium based on salt % ---
salt_pct = 1.0  # 1% salt
natural_na = rk.get('307', 0.0)
added_na = salt_pct * 393  # mg per 100g for 1% salt
rk['307'] = round(natural_na + added_na, 2)

rk1= pd.DataFrame(rk).T.assign(
    Long_Desc='Red cabbage sauerkraut (1% salt, inferred)',
    price=0.45,
    date_modified=date.today().isoformat()
).rename({0: red_kraut_id})

chosenfoods1 = pd.concat([chosenfoods1,rk1])

# --- Step 6: Export to Excel with frozen header/ID ---
output_path = chosenfoods1.xlsx
chosenfoods1.to_excel(output_path, engine='openpyxl')

# Freeze row 1 and column A
wb = load_workbook(output_path)
ws = wb.active
ws.freeze_panes = ws['B2']
wb.save(output_path)

from datetime import date
import numpy as np
import pandas as pd

# Define IDs for your foods
carrot_id = '11124'        # Raw carrot (example NDB ID)
green_cabbage_id = '11109' # Green cabbage (raw)
sauerkraut_id = '11439'    # Sauerkraut (from green cabbage)
saurkarrot_id = '99998'    # New food ID for inferred fermented carrot

# Copy carrot nutrient profile
sk = npt.loc[carrot_id].astype(float).copy()

# Build transformation ratios from cabbage → kraut
gc = npt.loc[green_cabbage_id].astype(float)
gs = npt.loc[sauerkraut_id].astype(float)

# Use only nutrients both foods have
transform_nutrients = ['203', '204', '208', '269', '212', '291', '401']  # + add '255' if desired
existing_nutrients = [nutr for nutr in transform_nutrients if nutr in npt.columns]

# Avoid divide-by-zero
gc_safe = gc.replace(0, np.nan)
ratios = gs[existing_nutrients] / gc_safe[existing_nutrients]

# Apply transformation to carrot
for nutr in existing_nutrients:
    if pd.notna(ratios[nutr]) and nutr in sk:
        sk[nutr] = round(sk[nutr] * ratios[nutr], 3)

# Override sodium manually (1% salt = 1000 mg per 100g food)
sk['307'] = 1000

# Ensure all index and columns are strings
chosenfoods1.columns = chosenfoods1.columns.astype(str)
sk.index = sk.index.astype(str)

# Build 1-row DataFrame
new_row_df = pd.DataFrame(sk).T
new_row_df.index = [saurkarrot_id]
new_row_df.columns = new_row_df.columns.astype(str)

# Add metadata
new_row_df['Long_Desc'] = 'Fermented carrot (1% salt, inferred)'
new_row_df['price'] = 0.20
new_row_df['date_modified'] = date.today().isoformat()

# Append to chosenfoods1 with new columns if needed
chosenfoods1 = pd.concat([chosenfoods1, new_row_df], axis=0, sort=False)

# Redundant solving loops from pyomo.py
import pandas as pd
import gspread
import gspread_formatting
from gspread_formatting import format_cell_range, cellFormat, color
import gspread.utils
# Main loop (change to while True: for indefinite execution if desired)
for i in range(150):


# BEFORE SOLVING: Add constraints to exclude all previous diets
    if i > 0:
        for diet_num, previous_diet in enumerate(diet_history):
            # This is the key: we're using STORED VALUES from previous_diet
            # combined with VARIABLE REFERENCES from uspyomo.prob.bv

            # Create the no-good cut:
            # "At least one food must be different from this previous diet"

            # In C terms, this would be like:
            # int hamming_distance = 0;
            # for(j in all_foods) {
            #     if(previous_diet[j] == 1 && current_bv[j] == 0) hamming_distance++;
            #     if(previous_diet[j] == 0 && current_bv[j] == 1) hamming_distance++;
            # }
            # assert(hamming_distance >= 1);

            expr = sum(
                previous_diet[food_id] * (1 - uspyomo.prob.bv[food_id]) +  # Was ON, now OFF
                (1 - previous_diet[food_id]) * uspyomo.prob.bv[food_id]    # Was OFF, now ON
                for food_id in uspyomo.prob.bv
            )

            # Add this as a named constraint (important for debugging)
            constraint_name = f"no_repeat_diet_{diet_num}"
            if hasattr(uspyomo.prob, constraint_name):
                delattr(uspyomo.prob, constraint_name)  # Remove if exists

            setattr(uspyomo.prob, constraint_name, Constraint(expr=expr >= 1))

        print(f"\nIteration {i}: Excluding {len(diet_history)} previous diets")

    # NOW SOLVE with all the no-good cuts in place
    nzf2 = uspyomo.UScplex2(False)

    # AFTER SOLVING: Store this solution's VALUES (not variable references!)
    # This is like memcpy() in C - we're copying the actual values
    current_diet = {}
    selected_foods = []

    for food_id in uspyomo.prob.bv:
        # Dereference and store the actual value (0 or 1)
        food_is_selected = round(value(uspyomo.prob.bv[food_id]))
        current_diet[food_id] = food_is_selected

        if food_is_selected == 1:
            selected_foods.append(food_id)

    # Check for duplicates (this shouldn't happen if cuts are working)
    is_duplicate = False
    for prev_diet in diet_history:
        if all(current_diet[f] == prev_diet[f] for f in uspyomo.prob.bv):
            is_duplicate = True
            print(f"⚠️ WARNING: Diet {i} is a duplicate! The no-good cuts may not be working.")
            break

    if not is_duplicate:
        print(f"✓ Diet {i}: Found new combination with {len(selected_foods)} foods")
        # Show first few foods for verification
        print(f"  Sample foods: {selected_foods[:5]}...")

    # Add to history (this is our "database" of previous diets)
    diet_history.append(current_diet)



    fl = nzf2[nzf2['amounts'] > 1e-3][['Long_Desc', 'amounts']]
    flu = fl.merge(fl9, left_index=True, right_index=True, how='left')
    flu['M'] = flu['amounts'] / flu['Gm_Wgt']
    flu = fineprint(flu)  # add in the food drilldowns as %
    flu = flu.loc[:, ~flu.columns.str.endswith(('Hi', 'Lo'))]

    obj, st = uspyomo.slackscplex()
    col = ['-slacks cost', '+slacks cost']
    cols = ['nuta', '-slacks cost', '+slacks cost', 'min', 'max', '-slacks', '+slacks']

    st = st[~st.index.str.endswith(('Hi', 'Lo'))]
    st['NutrDesc'].fillna(st.index.to_series(), inplace=True)
    st1 = st.set_index('NutrDesc')
    st1 = st1[cols].transpose()
    flu = pd.concat([flu, st1])

    # Write flu to Excel files (if needed)
    flu.sort_values('amounts', ascending=False).to_excel(fluwriter, sheet_name=str(i), freeze_panes=(2, 2))
    st[cols].sort_values(col, ascending=False).to_excel(slackswriter, sheet_name=str(obj), freeze_panes=(1, 1))

    # Prepare flu for Google Sheets update (with headers)
    flu_with_headers = flu.fillna('').reset_index().transpose().reset_index().transpose()

    # Create a new worksheet in the Google Sheet (using the iteration number as title)
    worksheet = sheet.add_worksheet(str(i), flu.shape[0], flu.shape[1])

    # Freeze the first row and first two columns
    worksheet.freeze(rows=1, cols=2)

    # Determine the cell range for the update (starting at A1)
    num_rows, num_cols = flu_with_headers.shape
    cell_range = gspread.utils.rowcol_to_a1(1, 1) + ":" + gspread.utils.rowcol_to_a1(num_rows, num_cols)

    # Update the worksheet with the DataFrame values
    worksheet.update(cell_range, flu_with_headers.values.tolist())

    # (Optional) Add a note to cell A1 using the Sheets API service
    service.spreadsheets().batchUpdate(
        spreadsheetId=sheet.id,
        body={
            'requests': [{
                'updateCells': {
                    'range': {
                        'sheetId': worksheet.id,
                        'startRowIndex': 0,
                        'endRowIndex': 1,
                        'startColumnIndex': 0,
                        'endColumnIndex': 1
                    },
                    'rows': [{
                        'values': [{'note': str(value(uspyomo.prob.obj))}]
                    }],
                    'fields': 'note'
                }
            }]
        }
    ).execute()

    # (Optional) Adjust Excel column width for the Excel file, if applicable
    worksheet_ex = fluwriter.sheets[str(i)]
    worksheet_ex.column_dimensions['B'].width = 50

    # -------------------------
    # Bi-Directional Highlighting and In-Loop Summary Update
    # -------------------------
    # Locate the "Long_Desc" column from the header row.
    header = worksheet.row_values(1)
    if "Long_Desc" not in header:
        raise ValueError('Column "Long_Desc" not found in header row.')
    col_index = header.index("Long_Desc") + 1  # Convert to 1-indexed

    # Get current "Long_Desc" values (skipping header)
    current_values = worksheet.col_values(col_index)[1:]

    # Retrieve the objective function value (as a string) for this iteration.
    # (Adjust the extraction as needed; here we assume value(uspyomo.prob.obj) provides the value.)
    obj_value = str(value(uspyomo.prob.obj))

    if i == 0:
        # For the first iteration, there's no previous sheet.
        summary_row = [str(i), obj_value, "\n"]
    else:
        # Get previous worksheet (by title, assumed to be str(i-1))
        prev_ws = sheet.worksheet(str(i - 1))
        previous_values = prev_ws.col_values(col_index)[1:]

        prev_set = set(previous_values)
        current_set = set(current_values)

        # Determine newly added and removed items.
        new_items = sorted(list(current_set - prev_set))
        removed_items = sorted(list(prev_set - current_set))#RWS fix this ? 2025-4-5
        print("New items:", new_items)
        print("Removed items:", removed_items)
        # # Use sets for order-independent comparison.
        # prev_set = set(previous_values)
        # current_set = set(current_values)

        # # Determine newly added and removed items.
        # new_items = sorted(list(current_set - prev_set))
        # removed_items = sorted(list(prev_set - current_set))#RWS fix this ? 2025-4-5

        # Highlight new items in current sheet (yellow)
        for row_num, val in enumerate(current_values, start=2):
            if val not in prev_set:
                cell_a1 = gspread.utils.rowcol_to_a1(row_num, col_index)
                fmt = cellFormat(backgroundColor=color(1, 1, 0))  # Yellow
                gspread_formatting.format_cell_range(worksheet, cell_a1, fmt)

        # Highlight removed items in previous sheet (cyan)
        for row_num, val in enumerate(previous_values, start=2):
            if val not in current_set:
                cell_a1 = gspread.utils.rowcol_to_a1(row_num, col_index)
                fmt = cellFormat(backgroundColor=color(0, 1, 1))  # Cyan
                gspread_formatting.format_cell_range(prev_ws, cell_a1, fmt)

        summary_row = [str(i), obj_value, "\n".join(new_items), "\n ".join(removed_items)]

    # Append the summary row directly to the "Summary" sheet.
    summary_ws.append_row(summary_row, value_input_option='USER_ENTERED')
    from gspread_formatting import format_cell_range, cellFormat
    wrap_fmt = cellFormat(wrapStrategy="WRAP")
    format_cell_range(summary_ws, "C:D", wrap_fmt)
    # Optional: Print log information
    print(f"Iteration {i} complete. Summary: {summary_row}")
